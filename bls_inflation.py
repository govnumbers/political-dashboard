#!/usr/bin/env python3
"""VA disability claims backlog, VA Monday Morning Workload Reports, weekly
XLSX at a predictable URL (benefits.va.gov/REPORTS/mmwr/{YYYY}/MMWR-MM-DD-YYYY.xlsx).

"Backlog" = rating claims pending more than 125 days (VA's own definition).
DEFINITION-GAMING GUARD (project doc 03): total pending is captured alongside,
because the backlog can fall while total pending doesn't.

The workbook's layout is label-based, not position-based, and this could not
be live-tested from the build sandbox, parser searches every sheet for the
labelled rows and raises on ambiguity (safe-fail). The inauguration baseline
is fetched live from the first weekly file of the term (week ending Sat
Jan 25, 2025) once, then reused from the stored series.

FIRST-RUN FIX (28 Jul 2026): files are named for the week-ending SATURDAY
(MMWR-07-04-2026.xlsx = Sat Jul 4), not the Monday they're posted, verified
against the live index. First run guessed Mondays and found nothing."""
import os
import sys
import io
import re
import datetime
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish, load_existing, UA  # noqa: E402

URL_FMT = "https://www.benefits.va.gov/REPORTS/mmwr/{y}/MMWR-{m:02d}-{d:02d}-{y}.{ext}"
BASELINE_DATE = datetime.date(2025, 1, 25)   # first week-ending Saturday of the term

# One-time weekly-archive backfill (phase 7): the MMWR files stay online back
# to ~2018 at the same URL pattern. Each daily run fills up to CHUNK missing
# week-ending Saturdays (newest first), so the full curve, including the 418k
# peak of Jan 2024, draws itself in over ~2 weeks of runs without ever making
# a single run heavy. Dates that 404 twice are remembered in the data file
# (archive_missing) and never re-tried; a backfill problem can never break the
# current week's publish.
ARCHIVE_START = datetime.date(2018, 1, 6)    # first week-ending Saturday of 2018
CHUNK = 30


def recent_saturdays(today, n=10):
    """Most recent week-ending Saturdays, newest first. Mon=0..Sat=5."""
    d = today - datetime.timedelta(days=(today.weekday() - 5) % 7)
    return [d - datetime.timedelta(weeks=i) for i in range(n)]


def saturdays_between(start, end):
    """Every week-ending Saturday in [start, end], ascending."""
    d = start + datetime.timedelta(days=(5 - start.weekday()) % 7)
    out = []
    while d <= end:
        out.append(d)
        d += datetime.timedelta(weeks=1)
    return out


def backfill_targets(existing, today, chunk=CHUNK):
    """Missing archive Saturdays to attempt this run: not already stored, not
    known-missing, newest first, capped at `chunk`."""
    have = {p["date"] for p in (existing or {}).get("series", [])}
    known_missing = set((existing or {}).get("archive_missing", []))
    latest = today - datetime.timedelta(days=7)   # current week handled by the live fetch
    todo = [d for d in saturdays_between(ARCHIVE_START, latest)
            if d.isoformat() not in have and d.isoformat() not in known_missing]
    return sorted(todo, reverse=True)[:chunk]


def fetch_workbook_for(date):
    for ext in ("xlsx", "xlsm"):   # some vintages are macro-enabled .xlsm
        url = URL_FMT.format(y=date.year, m=date.month, d=date.day, ext=ext)
        r = requests.get(url, headers=UA, timeout=90)
        if r.status_code == 200 and r.content[:2] == b"PK":   # zip magic (xlsx/xlsm)
            return url, r.content
    return None, None


def _norm(c):
    return re.sub(r"\s+", " ", str(c)).strip().lower() if c is not None else ""


def _to_num(x):
    try:
        v = float(x)
        return int(round(v)) if v >= 0 else None
    except (TypeError, ValueError):
        return None


def counts_from_rows(rows):
    """(backlog, total_pending) from one sheet's rows, or None if this sheet
    doesn't carry the metric columns.

    Verified against the real 2026-07-25 file: 'Rating Bundle - SOJ' headers
    its metric columns '# Pending' / '# Pending > 125 Days' (cells contain
    literal newlines, so whitespace is normalised) and the national figure is
    the 'Compensation Total' row. The word 'backlog' no longer appears
    anywhere in the workbook, the >125-days column IS the official backlog
    definition, and the card's unit says so."""
    for hi, row in enumerate(rows[:15]):
        labels = [_norm(c) for c in row]
        bcol = next((i for i, l in enumerate(labels) if l.startswith("# pending > 125")), None)
        if bcol is None:
            continue
        tcol = next((i for i, l in enumerate(labels) if l == "# pending"), None)

        def row_rank(r):
            lbl = " ".join(_norm(c) for c in r[:bcol] if isinstance(c, str))
            if "compensation total" in lbl:
                return 0                      # the headline national row
            if "national" in lbl or lbl.endswith("total"):
                return 1
            return 2

        candidates = [(row_rank(r), idx, r)
                      for idx, r in enumerate(rows[hi + 1: hi + 60])
                      if bcol < len(r) and _to_num(r[bcol]) is not None]
        if not candidates:
            continue
        _, _, best = min(candidates, key=lambda t: (t[0], t[1]))
        backlog = _to_num(best[bcol])
        total = _to_num(best[tcol]) if (tcol is not None and tcol < len(best)) else None
        if backlog and backlog >= 500 and (total is None or backlog <= total):
            return backlog, total
    return None


def extract_counts(content):
    """(backlog, total_pending) from the workbook; prefers the Rating Bundle
    sheets (the official backlog basis), falls back to any sheet carrying the
    '# Pending > 125' columns (e.g. 'Transformation'). Raises if none do."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ordered = sorted(wb.worksheets,
                     key=lambda ws: 0 if "rating bundle" in ws.title.lower() else 1)
    for ws in ordered:
        got = counts_from_rows([list(r) for r in ws.iter_rows(values_only=True)])
        if got:
            return got
    raise RuntimeError("no '# Pending > 125' column with a plausible national row "
                       "found in VA MMWR workbook (layout changed?)")


def main():
    today = datetime.date.today()
    url = content = None
    for saturday in recent_saturdays(today):
        url, content = fetch_workbook_for(saturday)
        if content:
            file_date = saturday
            break
    if not content:
        raise RuntimeError("no recent VA Monday Morning Workload Report found "
                           "(tried the last 10 week-ending Saturdays, .xlsx and .xlsm)")

    backlog, total = extract_counts(content)

    # inauguration baseline: from stored history if we have it, else fetch once
    existing = load_existing("va_claims_backlog") or {}
    baseline = (existing.get("baseline") or {}).get("value")
    if baseline is None:
        _, base_content = fetch_workbook_for(BASELINE_DATE)
        if base_content:
            try:
                baseline, _ = extract_counts(base_content)
            except RuntimeError:
                baseline = None

    out = {
        "id": "va_claims_backlog", "name": "VA claims backlog",
        "category": "Health & Safety Net", "value": backlog, "unit": "claims >125 days",
        "as_of": file_date.isoformat(), "direction": "up_is_bad",
        "total_pending": total,
        "source": {"name": "Dept. of Veterans Affairs, Monday Morning Workload Report",
                   "url": "https://www.benefits.va.gov/reports/detailed_claims_data.asp"},
        "cadence": "Weekly",
        "note": "Disability-compensation rating claims pending more than 125 days (VA's backlog "
                "definition). Total pending shown for context, the backlog can fall while "
                "overall pending doesn't.",
    }
    if baseline is not None:
        out["baseline"] = {"label": "At inauguration (week ending Jan 25, 2025)", "value": baseline}

    # --- chunked archive backfill (never allowed to break the live publish) ---
    points = [{"date": file_date.isoformat(), "value": backlog}]
    missing = set((existing or {}).get("archive_missing", []))
    targets = backfill_targets(existing, today)
    filled = 0
    for d in targets:
        try:
            _, c = fetch_workbook_for(d)
            if not c:
                missing.add(d.isoformat())
                continue
            b, _t = extract_counts(c)
            points.append({"date": d.isoformat(), "value": b})
            filled += 1
        except Exception as e:                                    # noqa: BLE001
            missing.add(d.isoformat())
            print(f"  ! va_backlog: archive {d} unparseable ({e}), marked missing, moving on")
    if targets:
        print(f"  ✓ va_backlog: archive backfill {filled}/{len(targets)} weeks this run"
              f" ({len(missing)} permanently unavailable so far)")
    if missing:
        out["archive_missing"] = sorted(missing)

    publish(out, series=points)


if __name__ == "__main__":
    main()
