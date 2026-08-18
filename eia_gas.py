#!/usr/bin/env python3
"""Renewable share of US electricity generation, computed from Energy Information Administration's
Electric Power Monthly Table 1.1 (keyless XLSX), monthly.

v3 register #16 (Energy tab). A COMPUTED metric with the formula printed on
the card (precedent: effective tariff rate): renewable share = (conventional
hydro + utility-scale solar + all other renewables) ÷ total utility-scale
generation. Pumped-storage hydro is excluded (it's storage, net-negative);
small-scale rooftop solar is excluded and the card says so.

SOURCE MECHANICS: `eia.gov/electricity/monthly/xls/table_1_01.xlsx` is a
plain keyless download at a stable URL, refreshed monthly (~2-month lag).
VERIFIED against the real file (creator download, 13 Aug 2026): it's a compact
summary carrying ~current year + 2 prior years of monthly data (plus annual
totals back to 2016, which we don't use), NOT deep history; the card's chart
is short and accretes over time via merge-don't-overwrite. The parser is
LABEL-KEYED (exact column names) and dies loudly if the layout changes. 2024
annual cross-check: 22.7% (Energy Information Administration-reported ~23%)."""
import io
import os
import re
import sys
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish, UA  # noqa: E402

URL = "https://www.eia.gov/electricity/monthly/xls/table_1_01.xlsx"
TERM_START = "2025-01"
MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], 1)}

# REAL LAYOUT (verified against the file the creator downloaded, 13 Aug 2026):
# Energy Information Administration's Table 1.1 browser download is a COMPACT summary, not the by-fuel
# breakdown first assumed. Renewables are three exact columns , 
# 'Hydroelectric Conventional', 'Solar' (utility-scale), and 'Renewable
# Sources Excluding Hydroelectric and Solar' (wind + geothermal + biomass +
# wood), over 'Total Generation at Utility Scale Facilities'. Pumped storage
# and the 'Estimated ... Solar' small-scale columns are excluded by NOT being
# in the match set. Periods are month names under 'Year YYYY' section headers
# (September is written 'Sept'); 'Annual Totals', 'Year to Date' and 'Rolling
# 12 Months' sections carry bare years, which are ignored. Monthly coverage
# runs ~current year + 2 prior years (merge-don't-overwrite accretes history).
MONTHS["sept"] = 9  # Energy Information Administration abbreviates September
REN_COLS = {"hydroelectric conventional", "solar",
            "renewable sources excluding hydroelectric and solar"}
TOTAL_COL = "total generation at utility scale facilities"


def _norm(c):
    return re.sub(r"\s+", " ", str(c)).strip().lower() if c is not None else ""


def parse_table(rows):
    """Worksheet rows -> [{date: YYYY-MM, value: share%}] ascending."""
    hdr_i = next((i for i, r in enumerate(rows)
                  if any(_norm(c) == TOTAL_COL for c in r)), None)
    if hdr_i is None:
        raise RuntimeError("EPM table 1.1: header row not found (layout changed)")
    hdr = [_norm(c) for c in rows[hdr_i]]
    total_col = hdr.index(TOTAL_COL)
    ren_cols = [i for i, l in enumerate(hdr) if l in REN_COLS]
    if len(ren_cols) != 3:
        raise RuntimeError(f"EPM table 1.1: expected 3 renewable columns, found {len(ren_cols)} "
                           "(layout changed)")

    def _f(r, idx):
        try:
            v = r[idx]
            return float(str(v).replace(",", "")) if v not in (None, "", "--", "NM") else None
        except (TypeError, ValueError, IndexError):
            return None

    out, year = [], None
    for r in rows[hdr_i + 1:]:
        period = _norm(r[0] if r else None)
        ym = re.match(r"year (\d{4})$", period)          # 'Year 2025' section header
        if ym:
            year = int(ym.group(1))
            continue
        mon = MONTHS.get(period)                          # a month name under the current year
        if not mon or year is None:
            continue                                      # bare years, YTD/Rolling headers, footnotes
        total = _f(r, total_col)
        if not total or total <= 0:
            continue
        ren = sum(v for v in (_f(r, i) for i in ren_cols) if v is not None)
        out.append({"date": f"{year}-{mon:02d}", "value": round(ren / total * 100, 1)})
    if not out:
        raise RuntimeError("EPM table 1.1: parsed zero period rows (layout changed)")
    # de-dup (a month can appear in both a monthly section and nowhere else here,
    # but guard anyway); latest wins
    dedup = {}
    for p in out:
        dedup[p["date"]] = p["value"]
    return [{"date": d, "value": dedup[d]} for d in sorted(dedup)]


def main():
    from openpyxl import load_workbook
    r = requests.get(URL, headers=UA, timeout=120)
    r.raise_for_status()
    if r.content[:2] != b"PK":
        raise RuntimeError("EPM table 1.1: response is not an XLSX")
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    series = parse_table([list(row) for row in ws.iter_rows(values_only=True)])

    latest = series[-1]
    base = next((p["value"] for p in series if p["date"] == TERM_START), None)
    out = {
        "id": "renewable_share", "name": "Renewable share of electricity generation",
        "category": "Energy", "value": latest["value"], "unit": "%",
        "as_of": latest["date"], "direction": "neutral",
        "source": {"name": "Energy Information Administration Electric Power Monthly, Table 1.1",
                   "url": "https://www.eia.gov/electricity/monthly/"},
        "cadence": "Monthly", "stale_days": 90,
        "note": "Share of US utility-scale electricity generated from renewables "
                "(conventional hydro, wind, solar, geothermal, biomass), a computed "
                "ratio of Energy Information Administration's own generation-by-source figures. Excludes small-scale "
                "rooftop solar and pumped storage; the mix is seasonal (hydro peaks in "
                "spring, solar in summer), so same-month comparisons matter.",
    }
    if base is not None:
        out["baseline"] = {"label": "At inauguration (Jan 2025)", "value": base}
    publish(out, series=series)


if __name__ == "__main__":
    main()
