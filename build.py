#!/usr/bin/env python3
"""ICE detention, national average daily population (average daily population), fiscal-YTD, ICE.

SOURCE MECHANICS (learned from the real FY26 workbook, 28 Jul 2026): ICE now
posts DATED snapshots (e.g. FY26_detentionStats07202026.xlsx) instead of
republishing one fixed filename, the bare FY{yy}_detentionStats.xlsx URL now
404s, which made the old connector silently fall back to the FY25 file. The
connector therefore discovers the newest file in layers: (1) scrape the
landing page for detentionStats links; (2) probe recent dated URLs; (3) legacy
bare URLs as a last resort.

LAYOUT (FY26 vintage): per-facility average daily population lives in the 'Facilities FY*' sheet
under a two-tier header, a group header row ('FY26 average daily population: ...') with subheaders
below it (Level A–D, and Male/Female Crim/Non-Crim). National average daily population = the sum of
Level A–D across facility rows. The workbook carries the same total split two
independent ways, so the parser computes BOTH and requires agreement within 1%
,  a built-in integrity check (both = 62,517 on the Jul-20 snapshot; the old
parser's header guess summed a single subset column and got 37,831 on the same
file, and 26,767 live off the wrong file, hence this rewrite).

Also captured: the point-in-time 'Currently Detained' total (65,765 on Jul 20)
as an extra field for card context, the headline stays average daily population, as labelled, and
the data-through date from the Footnotes sheet becomes as_of."""
import os
import sys
import io
import re
import datetime
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish, UA  # noqa: E402

LANDING = "https://www.ice.gov/detain/detention-management"
DOCLIB = "https://www.ice.gov/doclib/detention/"
BARE = DOCLIB + "FY{yy:02d}_detentionStats.xlsx"
DATED = DOCLIB + "FY{yy:02d}_detentionStats{mdY}.xlsx"


def current_fy(d):
    return d.year + 1 if d.month >= 10 else d.year


def _get(url):
    try:
        r = requests.get(url, headers=UA, timeout=90)
        if r.status_code == 200 and r.content[:2] == b"PK" and len(r.content) > 5000:
            return r.content
    except requests.RequestException:
        pass
    return None


def fetch_workbook():
    """Locate and download the newest detention-stats workbook.
    Returns (fy, url, content)."""
    today = datetime.date.today()
    fy = current_fy(today)

    # 1) landing page: hrefs like .../FY26_detentionStats07202026.xlsx
    try:
        page = requests.get(LANDING, headers=UA, timeout=60).text
        cands = []
        for m in re.finditer(r'href="([^"]*detentionStats[^"]*\.xlsx)"', page, re.I):
            url = m.group(1)
            if url.startswith("/"):
                url = "https://www.ice.gov" + url
            dm = re.search(r"(\d{2})(\d{2})(\d{4})\.xlsx$", url)          # MMDDYYYY suffix
            sort_key = f"{dm.group(3)}{dm.group(1)}{dm.group(2)}" if dm else "0"
            fm = re.search(r"FY(\d{2})_", url, re.I)
            cfy = int(fm.group(1)) + 2000 if fm else fy
            cands.append((sort_key, cfy, url))
        for _, cfy, url in sorted(cands, reverse=True):
            content = _get(url)
            if content:
                return cfy, url, content
    except requests.RequestException:
        pass

    # 2) probe recent dated URLs (posting day varies; last 35 days, newest first)
    fy_list = [fy] + ([fy - 1] if today.month in (10, 11) else [])
    for cand_fy in fy_list:
        for back in range(35):
            d = today - datetime.timedelta(days=back)
            content = _get(DATED.format(yy=cand_fy % 100, mdY=d.strftime("%m%d%Y")))
            if content:
                return cand_fy, DATED.format(yy=cand_fy % 100, mdY=d.strftime("%m%d%Y")), content

    # 3) legacy bare filenames (pre-Jul-2026 behaviour)
    for cand_fy in (fy, fy - 1):
        content = _get(BARE.format(yy=cand_fy % 100))
        if content:
            return cand_fy, BARE.format(yy=cand_fy % 100), content

    raise RuntimeError("could not locate any ICE detention workbook "
                       "(landing-page scrape + dated probes + legacy URLs all failed)")


def adp_from_rows(rows):
    """rows = list of row-lists from a Facilities sheet. Returns (adp, n_facilities).
    Sums Level A–D per-facility average daily population; cross-checks against the Male/Female
    Crim/Non-Crim split when present (the two must agree within 1%)."""
    sub_idx = None
    for i, r in enumerate(rows[:15]):
        labels = [str(c).strip().lower() if c is not None else "" for c in r]
        if "level a" in labels:
            sub_idx = i
            break
    if sub_idx is None:
        raise RuntimeError("Facilities sheet: 'Level A' subheader not found (layout changed)")

    labels = [str(c).strip().lower() if c is not None else "" for c in rows[sub_idx]]
    lvl = [i for i, l in enumerate(labels) if re.fullmatch(r"level [a-d]", l)]
    crim = [i for i, l in enumerate(labels)
            if l in ("male crim", "male non-crim", "female crim", "female non-crim")]
    if len(lvl) != 4:
        raise RuntimeError(f"Facilities sheet: expected 4 'Level A-D' columns, found {len(lvl)}")

    def colsum(cols):
        tot, n = 0.0, 0
        for r in rows[sub_idx + 1:]:
            name = str(r[0]).strip().lower() if r and r[0] is not None else ""
            if not name or "total" in name:
                continue
            counted = False
            for i in cols:
                if i < len(r):
                    try:
                        tot += float(r[i])
                        counted = True
                    except (TypeError, ValueError):
                        continue
            n += counted
        return tot, n

    adp, n = colsum(lvl)
    if adp <= 0 or n < 20:
        raise RuntimeError(f"Facilities sheet: implausible average daily population sum ({adp:.0f} from {n} facilities)")
    if len(crim) == 4:
        alt, _ = colsum(crim)
        if alt > 0 and abs(alt - adp) / adp > 0.01:
            raise RuntimeError(f"average daily population cross-check failed: Level-split {adp:.0f} vs "
                               f"criminality-split {alt:.0f}, treating as bad parse")
    return round(adp), n


def currently_detained(wb):
    """Point-in-time detained total from the 'Detention FY*' sheet (best-effort)."""
    det = next((ws for ws in wb.worksheets
                if re.match(r"\s*detention fy", ws.title, re.I)), None)
    if det is None:
        return None
    rows = [list(r) for r in det.iter_rows(values_only=True)]
    for ri, r in enumerate(rows[:25]):
        for ci, c in enumerate(r):
            if c is None or "currently detained by processing" not in str(c).lower():
                continue
            try:
                hdr = [str(x).strip().lower() if x is not None else "" for x in rows[ri + 1]]
                tcol = hdr.index("total", ci)
                for rr in rows[ri + 2: ri + 8]:
                    if rr and str(rr[ci]).strip().lower() == "total":
                        return int(float(rr[tcol]))
            except (ValueError, IndexError, TypeError):
                return None
    return None


def through_date(wb, kind):
    """Newest 'updated through MM/DD/YYYY' footnote date for `kind`
    ('Detention' / 'Removals'), as ISO, or None."""
    fn = next((ws for ws in wb.worksheets if ws.title.strip().lower() == "footnotes"), None)
    if fn is None:
        return None
    text = "\n".join(str(c) for r in fn.iter_rows(values_only=True)
                     for c in r if c is not None)
    dates = []
    for m in re.findall(rf"{kind}[^\n]*?through (\d{{2}}/\d{{2}}/\d{{4}})", text, re.I):
        try:
            dates.append(datetime.datetime.strptime(m, "%m/%d/%Y").date())
        except ValueError:
            continue
    return max(dates).isoformat() if dates else None


def main():
    from openpyxl import load_workbook
    fy, url, content = fetch_workbook()
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    fac = next((ws for ws in wb.worksheets if "facilities" in ws.title.lower()), None)
    if fac is None:
        raise RuntimeError("no 'Facilities' sheet in ICE workbook")
    adp, n_fac = adp_from_rows([list(r) for r in fac.iter_rows(values_only=True)])

    as_of = through_date(wb, "Detention") or datetime.date.today().isoformat()
    out = {
        "id": "ice_detention", "name": "ICE detention population (average daily population)",
        "category": "Immigration", "value": adp, "unit": "avg daily", "as_of": as_of,
        "direction": "neutral",
        "source": {"name": "U.S. Immigration and Customs Enforcement", "url": LANDING},
        "cadence": "Biweekly", "stale_days": 75,
        "note": f"National average daily population in ICE detention, fiscal-YTD (FY{fy}), "
                f"summed across {n_fac} facilities and cross-checked against the workbook's "
                "second split. An average, not a point-in-time headcount.",
    }
    cd = currently_detained(wb)
    if cd:
        out["currently_detained"] = cd

    # v3 lock: verified annual average daily population backfill (FY2019-24, ICE's own reports , 
    # citations + component checks inside the static file). Enhancement-only.
    try:
        import json as _json
        with open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "static", "ice_adp_annual.json")) as f:
            st = _json.load(f)
        out["annual_adp"] = {"values": st["adp_by_fy"], "source_note": st["_source"]}
    except Exception as e:  # noqa: BLE001
        print(f"  ! ice_detention: annual average daily population backfill skipped this run ({e})")

    publish(out, series=[{"date": as_of[:7], "value": adp}])


if __name__ == "__main__":
    main()
