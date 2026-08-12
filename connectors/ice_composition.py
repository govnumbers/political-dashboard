#!/usr/bin/env python3
"""ICE detention criminality composition — share of detainees with NO criminal
conviction, from ICE's own workbook. v3 register #20.

THE SAME FILE, A THIRD METRIC: reuses ice_detention's workbook discovery (one
download already feeds detention ADP + removals; this adds zero fetch risk).
The 'Detention FY*' sheet carries the detained population by ICE's own
criminality categories — verbatim labels "Convicted Criminal", "Pending
Criminal Charges", "Other Immigration Violator" — split by arresting agency.
TRAC and others compute the published no-conviction share from exactly these
fields (70.6% of 65,765 at the Jul 11 2026 snapshot).

INTEGRITY CHECK (same philosophy as the ADP dual-split): the three category
totals must sum to within 2% of the workbook's own point-in-time 'Currently
Detained' total, or the parse is treated as wrong and the run fails loudly.

Headline = % of currently detained with no criminal conviction (pending
charges + other immigration violators, both shown). 'No conviction' includes
people with charges pending — the split is stored and displayed, never
blurred. Labelled everywhere as ICE's own categories."""
import io
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish  # noqa: E402
from ice_detention import fetch_workbook, through_date, currently_detained, LANDING  # noqa: E402

LABELS = {
    "convicted": re.compile(r"^convicted criminal", re.I),
    "pending": re.compile(r"^pending criminal charge", re.I),
    "other": re.compile(r"^other immigration violator", re.I),
}


def criminality_block(rows):
    """Rows of the 'Detention FY*' sheet -> {convicted, pending, other} counts.

    Label-anchored: find each category row by its verbatim label in the first
    few columns, then take the row's RIGHTMOST plausible numeric cell (the
    workbook's blocks end in a Total column; agency sub-columns sit left of
    it). If a label appears in multiple blocks (ADP-based vs point-in-time),
    prefer the occurrence in the block nearest the 'Currently Detained'
    anchor; in practice the point-in-time block is the one whose three
    categories reconcile to the Currently Detained total — the caller's
    integrity check settles it."""
    found = {}
    for ri, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        for key, pat in LABELS.items():
            for ci, cell in enumerate(cells[:6]):
                if pat.match(cell):
                    nums = []
                    for c in r[ci + 1:]:
                        try:
                            v = float(str(c).replace(",", ""))
                            if v >= 0:
                                nums.append(v)
                        except (TypeError, ValueError):
                            continue
                    if nums:
                        found.setdefault(key, []).append((ri, nums[-1]))
    missing = [k for k in LABELS if k not in found]
    if missing:
        raise RuntimeError(f"criminality labels not found in Detention sheet: {missing} (layout changed)")
    return found


def reconcile(found, reference_total, tol=0.02):
    """Pick one occurrence per category such that the three sum to within
    `tol` of reference_total. Tries same-block (row-adjacent) combinations
    first. Returns {key: int} or raises."""
    import itertools
    combos = sorted(
        itertools.product(found["convicted"], found["pending"], found["other"]),
        key=lambda c: max(x[0] for x in c) - min(x[0] for x in c))  # tightest row-span first
    for (rc, vc), (rp, vp), (ro, vo) in combos:
        total = vc + vp + vo
        if reference_total and total > 0 and abs(total - reference_total) / reference_total <= tol:
            return {"convicted": int(vc), "pending": int(vp), "other": int(vo)}
    # no reference (workbook variant without the anchor): accept the tightest
    # row-adjacent trio if the three rows sit within 6 rows of each other
    (rc, vc), (rp, vp), (ro, vo) = combos[0]
    if max(rc, rp, ro) - min(rc, rp, ro) <= 6 and not reference_total:
        return {"convicted": int(vc), "pending": int(vp), "other": int(vo)}
    raise RuntimeError("criminality categories found but could not be reconciled "
                       f"to the Currently Detained total ({reference_total}) — treating as bad parse")


def main():
    from openpyxl import load_workbook
    fy, url, content = fetch_workbook()
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    det = next((ws for ws in wb.worksheets if re.match(r"\s*detention fy", ws.title, re.I)), None)
    if det is None:
        raise RuntimeError("no 'Detention FY*' sheet in ICE workbook")
    rows = [list(r) for r in det.iter_rows(values_only=True)]

    cd_total = currently_detained(wb)
    cats = reconcile(criminality_block(rows), cd_total)
    total = sum(cats.values())
    no_conviction = cats["pending"] + cats["other"]
    share = round(no_conviction / total * 100, 1)

    as_of = through_date(wb, "Detention") or None
    out = {
        "id": "ice_composition", "name": "ICE detainees with no criminal conviction",
        "category": "Immigration", "value": share, "unit": "%",
        "as_of": as_of or __import__("datetime").date.today().isoformat(),
        "direction": "neutral",
        "detail": {"convicted_criminal": cats["convicted"],
                   "pending_criminal_charges": cats["pending"],
                   "other_immigration_violators": cats["other"],
                   "total_detained": total},
        "source": {"name": "U.S. Immigration and Customs Enforcement (detention workbook)",
                   "url": LANDING},
        "cadence": "Biweekly", "stale_days": 75,
        "note": f"Share of the {total:,} people in ICE detention with no criminal "
                "conviction, using ICE's own categories: 'Pending Criminal Charges' plus "
                "'Other Immigration Violators' (no conviction and no pending charge), vs "
                "'Convicted Criminal'. The three categories are ICE's, reconciled against "
                "the workbook's own Currently Detained total before publishing.",
    }
    publish(out, series=[{"date": (as_of or out["as_of"])[:7], "value": share}])


if __name__ == "__main__":
    main()
