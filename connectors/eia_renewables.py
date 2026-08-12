#!/usr/bin/env python3
"""Renewable share of US electricity generation — computed from EIA's
Electric Power Monthly Table 1.1 (keyless XLSX), monthly, history to 2001.

v3 register #16 (Energy tab). A COMPUTED metric with the formula printed on
the card (precedent: effective tariff rate): renewable share = (conventional
hydro + wind + solar + geothermal + wood + other biomass) ÷ total generation,
UTILITY-SCALE only, monthly. Pumped-storage hydro is excluded (it's storage,
net-negative); small-scale rooftop solar is excluded and the card says so.

SOURCE MECHANICS: `eia.gov/electricity/monthly/xls/table_1_01.xlsx` is a
plain keyless download at a stable URL (verified Aug 2026), refreshed monthly
(~2-month lag). The sheet is a header block followed by one row per period.
The parser is LABEL-KEYED — it finds the header row by its column names and
matches source columns by keyword, so column reordering survives; a layout
change it can't resolve dies loudly on the validators instead of guessing.
First-run watchlist item (same playbook that shipped ICE/VA)."""
import io
import os
import re
import sys
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish, UA  # noqa: E402

URL = "https://www.eia.gov/electricity/monthly/xls/table_1_01.xlsx"
TERM_START = "2025-01"
MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], 1)}

RENEWABLE_PATTERNS = [
    r"hydroelectric conventional|conventional hydroelectric",
    r"\bwind\b",
    r"solar",              # matches photovoltaic + thermal utility columns
    r"geothermal",
    r"wood",
    r"other biomass|waste biomass|biomass",
]
EXCLUDE_PATTERN = re.compile(r"pumped storage|small[- ]scale", re.I)


def _norm(c):
    return re.sub(r"\s+", " ", str(c)).strip().lower() if c is not None else ""


def parse_table(rows):
    """Worksheet rows -> [{date: YYYY-MM, value: share%}] ascending.
    rows = list of row-lists (values only)."""
    # 1) header row: contains 'total' plus at least wind and solar labels
    hdr_i, hdr = None, None
    for i, r in enumerate(rows[:40]):
        labels = [_norm(c) for c in r]
        if any("wind" in l for l in labels) and any(l == "total" or l.endswith(" total") for l in labels):
            hdr_i, hdr = i, labels
            break
    if hdr is None:
        raise RuntimeError("EPM table 1.1: header row not found (layout changed)")

    total_col = next(i for i, l in enumerate(hdr) if l == "total" or l.endswith(" total"))
    ren_cols = []
    for i, l in enumerate(hdr):
        if not l or EXCLUDE_PATTERN.search(l):
            continue
        if any(re.search(p, l) for p in RENEWABLE_PATTERNS):
            ren_cols.append(i)
    if len(ren_cols) < 4:
        raise RuntimeError(f"EPM table 1.1: only {len(ren_cols)} renewable columns matched (layout changed)")

    # 2) period rows: EPM writes 'YYYY Month' in the first column (or Year, Month split)
    out = []
    for r in rows[hdr_i + 1:]:
        period = _norm(r[0] if r else None)
        m = re.match(r"^(19|20)(\d{2})\s+([a-z]+)", period)
        ym = None
        if m:
            mon = MONTHS.get(m.group(3))
            if mon:
                ym = f"{m.group(1)}{m.group(2)}-{mon:02d}"
        elif len(r) > 1:
            y, mo = _norm(r[0]), _norm(r[1])
            if re.fullmatch(r"(19|20)\d{2}", y) and mo in MONTHS:
                ym = f"{y}-{MONTHS[mo]:02d}"
        if not ym:
            continue

        def _f(idx):
            try:
                v = r[idx]
                return float(str(v).replace(",", "")) if v not in (None, "", "--", "NM") else None
            except (TypeError, ValueError, IndexError):
                return None

        total = _f(total_col)
        if not total or total <= 0:
            continue
        ren = sum(v for v in (_f(i) for i in ren_cols) if v is not None)
        out.append({"date": ym, "value": round(ren / total * 100, 1)})
    if not out:
        raise RuntimeError("EPM table 1.1: parsed zero period rows (layout changed)")
    return sorted(out, key=lambda p: p["date"])


def main():
    from openpyxl import load_workbook
    r = requests.get(URL, headers=UA, timeout=120)
    r.raise_for_status()
    if r.content[:2] != b"PK":
        raise RuntimeError("EPM table 1.1: response is not an XLSX")
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    series = parse_table([list(row) for row in ws.iter_rows(values_only=True)])

    latest = series[-1]
    base = next((p["value"] for p in series if p["date"] == TERM_START), None)
    out = {
        "id": "renewable_share", "name": "Renewable share of electricity generation",
        "category": "Energy", "value": latest["value"], "unit": "%",
        "as_of": latest["date"], "direction": "neutral",
        "source": {"name": "EIA Electric Power Monthly, Table 1.1",
                   "url": "https://www.eia.gov/electricity/monthly/"},
        "cadence": "Monthly", "stale_days": 90,
        "note": "Share of US utility-scale electricity generated from renewables "
                "(conventional hydro, wind, solar, geothermal, biomass) — a computed "
                "ratio of EIA's own generation-by-source figures. Excludes small-scale "
                "rooftop solar and pumped storage; the mix is seasonal (hydro peaks in "
                "spring, solar in summer), so same-month comparisons matter.",
    }
    if base is not None:
        out["baseline"] = {"label": "At inauguration (Jan 2025)", "value": base}
    publish(out, series=series)


if __name__ == "__main__":
    main()
