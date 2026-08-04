#!/usr/bin/env python3
"""ICE removals — from the SAME workbook as ice_detention (one download serves
both connectors' fetch logic; each runs standalone).

LAYOUT (learned from the real FY26 workbook, 28 Jul 2026): removals live as a
small labelled block in the top-right of the 'Detention FY*' sheet —
"ICE Removals: FY2026" / "Total" = fiscal-YTD removals (356,389 on the Jul-20
snapshot, data through Jul 11), plus a family-unit (FAMU) sub-count (36,548).
One FYTD figure per snapshot; the stored series accumulates a monthly point
that updates in place, like ice_detention. The Footnotes sheet's
"Removals data are updated through MM/DD/YYYY" line becomes as_of.

LABELLING RULE (locked in project doc 03): this is "ICE's published workbook
figure" and is NEVER reconciled to press-release "deportation" totals, which
mix ICE removals with CBP actions counted differently.

Cross-president context: static FY2024 baseline from the ICE ERO FY2024 Annual
Report (271,484 removals) — the workbook series itself starts in FY2025.
Known source risk: ICE paused publication for 56 days in early 2026
(stale_days tolerates a normal gap but flags a prolonged one)."""
import os
import sys
import io
import re
import json
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish  # noqa: E402
from ice_detention import fetch_workbook, through_date  # noqa: E402

FY2024_REMOVALS = 271484   # ICE ERO FY2024 Annual Report (static comparator)
ANNUAL_STATIC = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "static", "ice_removals_annual.json")


def annual_history():
    """FY2012–FY2024 removals from ICE ERO annual reports — a committed static
    file (closed fiscal years never change). Missing file just skips the
    context series; the live metric is unaffected."""
    try:
        with open(ANNUAL_STATIC) as f:
            rows = json.load(f)["annual"]
        return sorted(({"fy": int(r["fy"]), "value": int(r["value"]), "source": r.get("source", "")}
                       for r in rows), key=lambda r: r["fy"])
    except Exception as e:                                        # noqa: BLE001
        print(f"  ! ice_removals: static annual history unavailable ({e})")
        return None


def _num(x):
    if x is None:
        return None
    try:
        return int(float(str(x).replace(",", "")))
    except ValueError:
        return None


def removals_from_rows(rows):
    """rows of a 'Detention FY*' sheet. Returns (fy_year, fytd_total, famu).
    Anchored on the literal "ICE Removals: FY____" heading; raises if absent."""
    for ri, r in enumerate(rows):
        for ci, c in enumerate(r):
            if c is None:
                continue
            m = re.match(r"\s*ICE Removals:\s*FY\s*(\d{4})", str(c), re.I)
            if not m:
                continue
            fy_year = int(m.group(1))
            total = famu = None
            for rr in rows[ri + 1: ri + 8]:
                seg = list(rr[max(0, ci - 1): ci + 7])
                lbl = next((str(x).strip().lower() for x in seg
                            if x is not None and str(x).strip()), "")
                nums = [n for n in (_num(x) for x in seg) if n is not None]
                if lbl == "total" and nums:
                    total = nums[0]
                elif "famu" in lbl and nums:
                    famu = nums[0]
            if total:
                return fy_year, total, famu
    raise RuntimeError("'ICE Removals: FY____' block not found (layout changed — inspect manually)")


def main():
    from openpyxl import load_workbook
    fy, url, content = fetch_workbook()
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    det = [ws for ws in wb.worksheets if re.match(r"\s*detention fy", ws.title, re.I)]
    fy_year = fytd = famu = None
    for ws in det or wb.worksheets:
        try:
            fy_year, fytd, famu = removals_from_rows(
                [list(r) for r in ws.iter_rows(values_only=True)])
            break
        except RuntimeError:
            continue
    if fytd is None:
        raise RuntimeError("could not locate removals data in the ICE workbook "
                           "(tab layout may have changed — inspect manually)")

    as_of = through_date(wb, "Removals") or datetime.date.today().isoformat()
    note = (f"Removals recorded in ICE's published statistics workbook, FY{fy_year} to date. "
            "Official workbook figure — not comparable to press-release 'deportation' totals, "
            "which mix in CBP actions.")
    if famu:
        note += f" Includes {famu:,} with a family-unit identifier."

    out = {
        "id": "ice_removals", "name": "ICE removals (fiscal-YTD)",
        "category": "Immigration", "value": fytd, "unit": "people", "as_of": as_of,
        "direction": "neutral",
        "comparison": {"label": "FY2024 full year (prior administration)", "value": FY2024_REMOVALS},
        "source": {"name": "U.S. Immigration and Customs Enforcement",
                   "url": "https://www.ice.gov/detain/detention-management"},
        "cadence": "Biweekly", "stale_days": 75, "note": note,
    }
    if famu:
        out["famu_removals"] = famu
    hist = annual_history()
    if hist:
        out["annual_history"] = hist
    publish(out, series=[{"date": as_of[:7], "value": fytd}])


if __name__ == "__main__":
    main()
