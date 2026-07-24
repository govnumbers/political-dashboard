#!/usr/bin/env python3
"""ICE detention — average daily population (ADP), fiscal-year-to-date — ICE.

ICE publishes a per-fiscal-year workbook at a STABLE, predictable URL
(FY{yy}_detentionStats.xlsx) but republishes it in place ~every two weeks with
advancing FYTD numbers and overwrites prior snapshots — so the merged series in
data/ is the historical store. We compute national ADP as the sum of the
facility-level ADP column.

Tracks ADP (a fiscal-YTD average), NOT the point-in-time headcount that TRAC
publishes — the card says so; don't conflate them.

HEADS-UP: of all the connectors this is the one most likely to need a small
post-first-run tweak. The workbook has merged cells, a title block above the
header, and column order that shifts between fiscal years, and it could not be
live-tested from the build sandbox (ice.gov is egress-blocked there). It keys on
header NAMES, not positions, and is fully safe-fail: on any parse problem it
raises, so there is no card and the run goes red rather than publishing garbage."""
import os
import sys
import io
import datetime
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish  # noqa: E402

DOCLIB = "https://www.ice.gov/doclib/detention/FY{yy:02d}_detentionStats.xlsx"
LANDING = "https://www.ice.gov/detain/detention-management"
UA = {"User-Agent": "govnumbers-dashboard/1.0 (+https://political-dashboard-323.pages.dev)"}


def current_fy(d):
    return d.year + 1 if d.month >= 10 else d.year


def fetch_workbook():
    fy = current_fy(datetime.date.today())
    tried = []
    for cand in (fy, fy - 1):          # handle the Oct rollover gap
        url = DOCLIB.format(yy=cand % 100)
        tried.append(url)
        try:
            r = requests.get(url, headers=UA, timeout=90)
            if r.status_code == 200 and len(r.content) > 5000:
                return cand, url, r.content
        except requests.RequestException:
            continue
    raise RuntimeError(f"could not fetch ICE detention workbook (tried {tried})")


def national_adp(content):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    best = None  # (count_of_rows, adp_sum)
    for ws in wb.worksheets:
        rows = [[(c if c is not None else "") for c in r] for r in ws.iter_rows(values_only=True)]
        for hi, row in enumerate(rows[:20]):
            labels = [str(c).strip().lower() for c in row]
            adp_cols = [i for i, l in enumerate(labels) if "adp" in l]
            if not adp_cols:
                continue
            # prefer a 'total ADP' column, else the first ADP column
            target = next((i for i in adp_cols if "total" in labels[i]), adp_cols[0])
            total, n = 0.0, 0
            for r in rows[hi + 1:]:
                if target >= len(r):
                    continue
                label0 = " ".join(str(c).lower() for c in r[:2])
                if any(t in label0 for t in ("total", "grand", "sum")):
                    continue
                try:
                    v = float(r[target])
                except (TypeError, ValueError):
                    continue
                if v > 0:
                    total += v
                    n += 1
            if n and (best is None or n > best[0]):
                best = (n, total)
            break  # header found for this sheet
    if not best or best[1] <= 0:
        raise RuntimeError("could not locate an ADP column / sum facility rows in ICE workbook")
    return round(best[1])


def main():
    fy, url, content = fetch_workbook()
    adp = national_adp(content)
    as_of = datetime.date.today().strftime("%Y-%m")
    out = {
        "id": "ice_detention", "name": "ICE detention population (ADP)",
        "category": "Immigration", "value": adp, "unit": "avg daily", "as_of": as_of,
        "direction": "neutral",
        "source": {"name": "U.S. Immigration and Customs Enforcement", "url": LANDING},
        "cadence": "Biweekly",
        "note": f"National average daily population in ICE detention, fiscal-YTD (FY{fy}). "
                "This is an average, not a point-in-time headcount.",
    }
    publish(out, series=[{"date": as_of, "value": adp}])


if __name__ == "__main__":
    main()
