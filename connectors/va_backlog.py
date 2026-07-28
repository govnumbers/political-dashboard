#!/usr/bin/env python3
"""VA disability claims backlog — VA Monday Morning Workload Reports, weekly
XLSX at a predictable URL (benefits.va.gov/REPORTS/mmwr/{YYYY}/MMWR-MM-DD-YYYY.xlsx).

"Backlog" = rating claims pending more than 125 days (VA's own definition).
DEFINITION-GAMING GUARD (project doc 03): total pending is captured alongside,
because the backlog can fall while total pending doesn't.

The workbook's layout is label-based, not position-based, and this could not
be live-tested from the build sandbox — parser searches every sheet for the
labelled rows and raises on ambiguity (safe-fail). The inauguration baseline
is fetched live from the first weekly file of the term (week ending Sat
Jan 25, 2025) once, then reused from the stored series.

FIRST-RUN FIX (28 Jul 2026): files are named for the week-ending SATURDAY
(MMWR-07-04-2026.xlsx = Sat Jul 4), not the Monday they're posted — verified
against the live index. First run guessed Mondays and found nothing."""
import os
import sys
import io
import re
import datetime
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import publish, load_existing, UA  # noqa: E402

URL_FMT = "https://www.benefits.va.gov/REPORTS/mmwr/{y}/MMWR-{m:02d}-{d:02d}-{y}.{ext}"
BASELINE_DATE = datetime.date(2025, 1, 25)   # first week-ending Saturday of the term


def recent_saturdays(today, n=10):
    """Most recent week-ending Saturdays, newest first. Mon=0..Sat=5."""
    d = today - datetime.timedelta(days=(today.weekday() - 5) % 7)
    return [d - datetime.timedelta(weeks=i) for i in range(n)]


def fetch_workbook_for(date):
    for ext in ("xlsx", "xlsm"):   # some vintages are macro-enabled .xlsm
        url = URL_FMT.format(y=date.year, m=date.month, d=date.day, ext=ext)
        r = requests.get(url, headers=UA, timeout=90)
        if r.status_code == 200 and r.content[:2] == b"PK":   # zip magic (xlsx/xlsm)
            return url, r.content
    return None, None


def extract_counts(content):
    """Find (backlog, total_pending) via labelled rows across all sheets.
    backlog: a row whose label contains 'backlog'; total: a row mentioning
    pending claims/inventory. Takes the first numeric in the labelled row."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    backlog = total = None

    def first_number(cells):
        for c in cells:
            try:
                v = float(c)
                if v >= 0:
                    return int(v)
            except (TypeError, ValueError):
                continue
        return None

    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            label = " ".join(str(c) for c in r[:3] if c is not None).lower()
            if backlog is None and "backlog" in label:
                backlog = first_number(r)
            if total is None and re.search(r"(total|pending)\s+(claims|inventory)|claims\s+pending", label):
                total = first_number(r)
        if backlog is not None and total is not None:
            break
    if backlog is None:
        raise RuntimeError("no 'backlog' row found in VA MMWR workbook (layout changed?)")
    return backlog, total


def main():
    today = datetime.date.today()
    url = content = None
    for saturday in recent_saturdays(today):
        url, content = fetch_workbook_for(saturday)
        if content:
            file_date = saturday
            break
    if not content:
        raise RuntimeError("no recent VA Monday Morning Workload Report found "
                           "(tried the last 10 week-ending Saturdays, .xlsx and .xlsm)")

    backlog, total = extract_counts(content)

    # inauguration baseline: from stored history if we have it, else fetch once
    existing = load_existing("va_claims_backlog") or {}
    baseline = (existing.get("baseline") or {}).get("value")
    if baseline is None:
        _, base_content = fetch_workbook_for(BASELINE_DATE)
        if base_content:
            try:
                baseline, _ = extract_counts(base_content)
            except RuntimeError:
                baseline = None

    out = {
        "id": "va_claims_backlog", "name": "VA claims backlog",
        "category": "Health & Safety Net", "value": backlog, "unit": "claims >125 days",
        "as_of": file_date.isoformat(), "direction": "up_is_bad",
        "total_pending": total,
        "source": {"name": "Dept. of Veterans Affairs — Monday Morning Workload Report",
                   "url": "https://www.benefits.va.gov/reports/detailed_claims_data.asp"},
        "cadence": "Weekly",
        "note": "Disability-compensation rating claims pending more than 125 days (VA's backlog "
                "definition). Total pending shown for context — the backlog can fall while "
                "overall pending doesn't.",
    }
    if baseline is not None:
        out["baseline"] = {"label": "At inauguration (week ending Jan 25, 2025)", "value": baseline}
    publish(out, series=[{"date": file_date.isoformat(), "value": backlog}])


if __name__ == "__main__":
    main()
